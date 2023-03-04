# Excel Utility file contains reusable functions

import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file,sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return (sheet.max_row)

def getColumnCount(file,sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return (sheet.max_column)

def readData(file,sheet_name,rowNum,coluNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(rowNum,coluNum).value

def writeData(file,sheet_name,rowNum,columNum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(rowNum,columNum).value = data
    workbook.save(file)

def fillGreenColor(file,sheet_name,rowNum,columNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rowNum,columNum).fill = greenFill
    workbook.save(file)

def fillRedColor(file,sheet_name,rowNum,columNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    redFill = PatternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='solid')
    sheet.cell(rowNum,columNum).fill = redFill
    workbook.save(file)